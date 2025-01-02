import openpyxl  # type: ignore
import re
from collections import defaultdict

log_file = "merge_log_he10_real.txt"
id_list = [3, 4, 5, 6, 7, 8, 9, 10]
log_files = [f"log_merging_he{i}.log" for i in id_list]
column_to_write = 7
# name_template = r"ps_(\d{6})_(\d{3}).root"
length_of_file_name = 18

# 打开 Excel 文件
workbook = openpyxl.load_workbook("template.xlsx")
worksheet = workbook['Sheet1']

# 使用 defaultdict 来确保键不存在时会初始化为 0
run_id_count = defaultdict(int)

# 读取并分析日志文件
for log_file in log_files:
    with open(log_file, "r") as file:
        for line in file.readlines():
            line = line.strip()
            if "hadd" in line:
                continue
            if "ps" in line and ".root" in line and len(line) == length_of_file_name:
                # 使用正则表达式找到所有六位数的数字
                run_ids = re.findall(r'_(\d{6})_', line)  # \b表示单词边界，\d{6}表示6位数字
                # print(run_ids)
                # 如果找到六位数的数字，可以选择第一个（或者所有找到的数字，根据需求）
                if len(run_ids) == 1:
                    run_id = run_ids[0]  # 这里选择第一个六位数数字作为 run_id
                    # 或者，若需要获取所有六位数的数字，直接使用 run_ids 列表d
                else :
                    exit("Match Error :" + line)
                # 更新字典中的 run_id 计数
                run_id_count[run_id] += 1

    # 输出字典内容（每个 run_id 对应的文件数量）
    for run_id, count in run_id_count.items():
        print(f"Run ID: {run_id}, Number of files: {count}")

    # 将统计结果写入 Excel 文件的第 8 列
    for i in range(1, worksheet.max_row + 1):
        run_id = worksheet.cell(i, 1).value  # 获取第 i 行的 run_id（假设它在第一列）
        
        # 只处理非空的 run_id
        if run_id in run_id_count:
            worksheet.cell(i, column_to_write, value=f"{run_id_count[run_id]}")  # 将计数结果写入第 8 列
        if str(run_id) in run_id_count:
            run_id = str(run_id)
            worksheet.cell(i, column_to_write, value=f"{run_id_count[run_id]}")  # 将计数结果写入第 8 列

# 保存修改后的 Excel 文件
workbook.save("template_modified.xlsx")
