import openpyxl  # type: ignore
from pprint import pprint
import os

run_id_column = 1
exclusion_note_column = 10
exclusion_text_file = "/work/halld/home/zhikun/lumi_skim/exclusion.log"
xlsx_file_template = "/work/halld/home/zhikun/lumi_skim/template_modified.xlsx"

# Read the exclusion log file and parse the lines
with open(exclusion_text_file, "r") as f:
    lines = [line.strip().split() for line in f.readlines()]

# Build the dictionary of bad runs and their associated files
bad_runs_and_files = {}
for line in lines:
    bad_runs_and_files[line[0]] = line[1:]

pprint(bad_runs_and_files)

# Load the Excel workbook
workbook = openpyxl.load_workbook(xlsx_file_template)
worksheet = workbook["Sheet1"]
max_row = worksheet.max_row

# Iterate over each row in the sheet and add exclusion notes if the run_id is in the bad runs dictionary
for i in range(1, max_row + 1):
    run_id = worksheet.cell(row=i, column=run_id_column).value
    run_id = str(run_id)
    if run_id in bad_runs_and_files:
        # Prepare the exclusion message
        exclusion_files = bad_runs_and_files[run_id]
        exclusion_msg = " ".join(exclusion_files)  # Join the list of file_ids into a single string
        print(exclusion_msg)
        # Write the exclusion message to the specified column
        worksheet.cell(row=i, column=exclusion_note_column, value=exclusion_msg)

# Save the updated workbook
output_file = "template_modified_with_exclusion_notes.xlsx"
workbook.save(output_file)

print(f"Updated Excel file saved as {output_file}")
